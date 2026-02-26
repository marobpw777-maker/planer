"""
ПРОФЕССИОНАЛЬНЫЙ ФИНАНСОВЫЙ КОМПАНЬОН 2026
С графическим интерфейсом и полной функциональностью Excel
Версия 8.0 - Исправленные критические ошибки + улучшения
"""

import tkinter as tk
from tkinter import ttk, messagebox
import pandas as pd
import openpyxl
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, GradientFill, NamedStyle
from openpyxl.formatting.rule import FormulaRule, ColorScaleRule, IconSetRule, DataBarRule
from openpyxl.chart import PieChart, LineChart, BarChart, Reference, ScatterChart, AreaChart, RadarChart
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.marker import DataPoint
from datetime import datetime, timedelta
from openpyxl.utils import get_column_letter
import warnings
import random
import math
from typing import List, Dict, Any, Optional
import textwrap
import sys
import os
import json
import shutil
from dateutil.relativedelta import relativedelta
warnings.filterwarnings('ignore')

class Config:
    """Конфигурация приложения"""
    MAX_CREDITS = 10
    MAX_INCOME_SOURCES = 20
    CURRENCY = '₽'
    DATE_FORMAT = '%d.%m.%Y'
    BACKUP_FOLDER = 'backups/'
    
    # Финансовые пороги
    CREDIT_LOAD_WARNING = 40  # % от дохода
    CREDIT_LOAD_CRITICAL = 60  # % от дохода
    FREE_CASH_WARNING = 10  # % от дохода
    FREE_CASH_CRITICAL = 5   # % от дохода

class BackupManager:
    """Менеджер резервного копирования"""
    
    @staticmethod
    def create_backup(filename):
        """Создание резервной копии файла"""
        try:
            if not os.path.exists(filename):
                return False
                
            if not os.path.exists(Config.BACKUP_FOLDER):
                os.makedirs(Config.BACKUP_FOLDER)
            
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            backup_name = f"backup_{timestamp}_{os.path.basename(filename)}"
            backup_path = os.path.join(Config.BACKUP_FOLDER, backup_name)
            
            shutil.copy2(filename, backup_path)
            return backup_path
        except Exception as e:
            print(f"Ошибка резервного копирования: {e}")
            return False
    
    @staticmethod
    def cleanup_old_backups(max_backups=10):
        """Очистка старых резервных копий"""
        try:
            if not os.path.exists(Config.BACKUP_FOLDER):
                return
            
            backups = []
            for file in os.listdir(Config.BACKUP_FOLDER):
                if file.startswith('backup_'):
                    file_path = os.path.join(Config.BACKUP_FOLDER, file)
                    backups.append((file_path, os.path.getmtime(file_path)))
            
            backups.sort(key=lambda x: x[1], reverse=True)
            
            for backup_path, _ in backups[max_backups:]:
                os.remove(backup_path)
        except Exception as e:
            print(f"Ошибка очистки бэкапов: {e}")

class UltimateFinancialCompanion:
    """Ультимативный финансовый компаньон с профессиональной визуализацией"""
    
    def __init__(self):
        self.credits = []
        self.incomes = []
        self.expenses = []
        self.goals = []
        self.today = datetime.now()
        self.wb = None
        self.setup_design_system()
        self.setup_behavioral_patterns()
        self.setup_chart_templates()
        
    def setup_design_system(self):
        """Профессиональная дизайн-система 2026"""
        self.colors = {
            'primary': {
                'dark_blue': 'FF0F2B46',
                'blue': 'FF1E5A8A',
                'light_blue': 'FF2E8BC0',
                'cyan': 'FF4FC3F7',
                'gradient_blue': ['FF0F2B46', 'FF1E5A8A', 'FF2E8BC0']
            },
            'accent': {
                'success': 'FF4CAF50',
                'success_light': 'FF81C784',
                'success_dark': 'FF2E7D32',
                'warning': 'FFFF9800',
                'warning_light': 'FFFFB74D',
                'warning_dark': 'FFEF6C00',
                'info': 'FF2196F3',
                'info_light': 'FF64B5F6',
                'info_dark': 'FF1565C0',
                'purple': 'FF9C27B0',
                'purple_light': 'FFBA68C8',
                'purple_dark': 'FF6A1B9A',
                'danger': 'FFF44336',
                'danger_light': 'FFEF5350',
                'danger_dark': 'FFC62828'
            },
            'background': {
                'light': 'FFF5F7FA',
                'lighter': 'FFFFFFFF',
                'dark': 'FF263238',
                'gray': 'FFECEFF1',
                'gray_light': 'FFF9FAFB'
            },
            'gradients': {
                'green': ['FF4CAF50', 'FF66BB6A', 'FF81C784'],
                'blue': ['FF2196F3', 'FF42A5F5', 'FF64B5F6'],
                'orange': ['FFFF9800', 'FFFFA726', 'FFFFB74D'],
                'purple': ['FF9C27B0', 'FFAB47BC', 'FFBA68C8'],
                'red': ['FFF44336', 'FFEF5350', 'FFE57373'],
                'cyan': ['FF00BCD4', 'FF26C6DA', 'FF4DD0E1']
            },
            'status': {
                'excellent': 'FF4CAF50',
                'good': 'FF8BC34A',
                'attention': 'FFFFC107',
                'critical': 'FFF44336',
                'neutral': 'FF9E9E9E',
                'info': 'FF2196F3'
            },
            'chart': {
                'bar_positive': 'FF4CAF50',
                'bar_negative': 'FFF44336',
                'bar_neutral': 'FFFFC107',
                'line_primary': 'FF2196F3',
                'line_secondary': 'FF9C27B0',
                'area_fill': 'FFE3F2FD',
                'pie_colors': ['FF4CAF50', 'FF2196F3', 'FFFF9800', 'FF9C27B0', 'FF00BCD4', 'FFFFC107']
            }
        }
        
        self.header_style = NamedStyle(name="header_style")
        self.header_style.font = Font(bold=True, color="FFFFFFFF", size=12)
        self.header_style.fill = PatternFill(
            start_color=self.colors['primary']['dark_blue'],
            end_color=self.colors['primary']['dark_blue'],
            fill_type="solid"
        )
        self.header_style.alignment = Alignment(horizontal="center", vertical="center")
        self.header_style.border = Border(
            left=Side(style='thin', color='FFFFFFFF'),
            right=Side(style='thin', color='FFFFFFFF'),
            top=Side(style='thin', color='FFFFFFFF'),
            bottom=Side(style='thin', color='FFFFFFFF')
        )
        
    def setup_behavioral_patterns(self):
        """Поведенческие паттерны и мотивационные системы"""
        self.positive_reinforcement = {
            'milestones': {
                10: "🎯 Первые 10% - отличное начало!",
                25: "🚀 Четверть пути пройдена!",
                50: "🎊 Половина сделана! Вы на финишной прямой!",
                75: "🌟 Осталось всего 25%!",
                90: "💫 Почти у цели! Последний рывок!",
                100: "🏆 Поздравляем! Цель достигнута!"
            }
        }
        
        self.financial_tips = [
            "💡 Автоматизируйте платежи - меньше шансов забыть",
            "💡 Сначала погашайте кредиты с самой высокой ставкой",
            "💡 Создайте подушку безопасности на 3-6 месяцев",
            "💡 Регулярно пересматривайте свои финансовые цели",
            "💡 Используйте кэшбэк и бонусы банков",
            "💡 Инвестируйте в свое финансовое образование",
            "💡 Празднуйте маленькие финансовые победы",
            "💡 Планируйте крупные покупки заранее",
            "💡 Сравнивайте предложения банков перед кредитом",
            "💡 Используйте приложения для отслеживания расходов"
        ]
        
    def setup_chart_templates(self):
        """Шаблоны для профессиональных диаграмм"""
        self.chart_templates = {
            'dashboard': {
                'width': 15,
                'height': 8,
                'title_font': Font(size=12, bold=True, color=self.colors['primary']['dark_blue']),
                'legend_position': 'b',
                'data_labels': True
            }
        }
        
    def get_tip_of_the_day(self):
        """Случайный финансовый совет"""
        return random.choice(self.financial_tips)
        
    def calculate_financial_health_score(self, credit):
        """Расчет комплексного балла финансового здоровья - УЛУЧШЕННАЯ ВЕРСИЯ"""
        if not credit or credit['amount'] <= 0:
            return {'total': 0, 'breakdown': {}, 'grade': 'Нет данных'}
        
        # 1. Прогресс погашения (макс 40 баллов)
        progress = (credit['paid'] / credit['amount']) * 100 if credit['amount'] > 0 else 0
        progress_score = min(progress * 0.4, 40)  # Более линейная зависимость
        
        # 2. Своевременность платежей (макс 30 баллов)
        months_passed = self.calculate_months_passed(credit)
        should_have_paid = credit.get('monthly_payment', 0) * months_passed
        
        if credit['paid'] >= should_have_paid:
            timeliness_score = 30
        elif credit['paid'] >= should_have_paid * 0.8:
            timeliness_score = 20
        elif credit['paid'] >= should_have_paid * 0.5:
            timeliness_score = 10
        else:
            timeliness_score = 0
        
        # 3. Оптимальность условий (макс 20 баллов)
        rate = credit['rate'] * 100  # Проценты
        
        if rate < 10:
            strategy_score = 20
        elif rate < 15:
            strategy_score = 15
        elif rate < 20:
            strategy_score = 10
        elif rate < 30:
            strategy_score = 5
        else:
            strategy_score = 0
        
        # 4. Соотношение платежа к доходу (макс 10 баллов)
        total_income = sum(i.get('amount', 0) for i in self.incomes) if self.incomes else 1
        payment_to_income = (credit.get('monthly_payment', 0) / total_income) * 100 if total_income > 0 else 100
        
        if payment_to_income < 10:
            comfort_score = 10
        elif payment_to_income < 20:
            comfort_score = 8
        elif payment_to_income < 30:
            comfort_score = 6
        elif payment_to_income < 40:
            comfort_score = 4
        elif payment_to_income < 50:
            comfort_score = 2
        else:
            comfort_score = 0
        
        total_score = progress_score + timeliness_score + strategy_score + comfort_score
        
        return {
            'total': round(total_score, 1),
            'breakdown': {
                'progress': round(progress_score, 1),
                'timeliness': round(timeliness_score, 1),
                'strategy': round(strategy_score, 1),
                'comfort': round(comfort_score, 1)
            },
            'grade': self.get_health_grade(total_score)
        }
        
    def get_health_grade(self, score):
        """Оценка финансового здоровья"""
        if score >= 90:
            return "🏆 Отличное"
        elif score >= 75:
            return "✅ Хорошее"
        elif score >= 60:
            return "⚠️ Требует внимания"
        else:
            return "🚨 Критическое"
            
    def calculate_monthly_payment(self, credit):
        """Расчет аннуитетного платежа"""
        if isinstance(credit, dict):
            amount = credit.get('amount', 0)
            rate = credit.get('rate', 0)
            months = credit.get('months', 1)
        else:
            amount = credit
            rate = 0.1
            months = 12
            
        if rate > 0:
            monthly_rate = rate / 12
            if monthly_rate > 0 and months > 0:
                monthly_payment = amount * (monthly_rate * (1 + monthly_rate) ** months) / ((1 + monthly_rate) ** months - 1)
            else:
                monthly_payment = amount / months if months > 0 else 0
        else:
            monthly_payment = amount / months if months > 0 else 0
        return monthly_payment
        
    def calculate_months_passed(self, credit):
        """Расчет прошедших месяцев"""
        if 'start_date' not in credit:
            return 0
            
        if isinstance(credit['start_date'], str):
            try:
                start_date = datetime.strptime(credit['start_date'], '%d.%m.%Y')
            except:
                start_date = datetime.now()
        else:
            start_date = credit['start_date']
            
        months_passed = (self.today.year - start_date.year) * 12 + \
                       (self.today.month - start_date.month)
        if self.today.day < start_date.day:
            months_passed -= 1
        return max(0, min(months_passed, credit.get('months', 0)))
    
    def create_excel_file(self):
        """Создание профессионального Excel-файла с динамическими формулами"""
        try:
            # Проверка данных перед созданием
            if not any([self.credits, self.incomes, self.expenses, self.goals]):
                print("⚠️ Нет данных для создания отчета")
                return False
            
            self.wb = Workbook()
            
            # Удаляем дефолтный лист
            if 'Sheet' in self.wb.sheetnames:
                default_sheet = self.wb['Sheet']
                self.wb.remove(default_sheet)
                
            print("   СОЗДАНИЕ ФИНАНСОВОГО КОМПАНЬОНА...")
            
            # Порядок листов
            self.create_summary_sheet()
            if self.incomes:
                self.create_income_sheet()
            if self.credits:
                self.create_credit_sheets()
            if self.expenses:
                self.create_expense_sheet()
            if self.goals:
                self.create_goals_sheet()
            self.create_analytics_sheet()
            self.create_calendar_sheet()
            self.create_instructions_sheet()
            self.create_dashboard_sheet()
            
            # Добавляем визуальные улучшения
            self.complete_visual_upgrade()
            
            # Сохраняем файл
            filename = f"Финансовый_Компаньон_Профессионал_{self.today.strftime('%Y%m%d_%H%M')}.xlsx"
            self.wb.save(filename)
            
            # Создаем резервную копию
            BackupManager.create_backup(filename)
            BackupManager.cleanup_old_backups()
            
            print("   ✅ Файл успешно создан!")
            return filename
            
        except Exception as e:
            print(f"❌ Критическая ошибка при создании файла: {e}")
            import traceback
            error_details = traceback.format_exc()
            
            # Записать ошибку в лог
            with open("error_log.txt", "a", encoding="utf-8") as f:
                f.write(f"{datetime.now()} - Ошибка: {str(e)}\n")
                f.write(f"Детали: {error_details}\n")
            
            return False
    
    def create_dashboard_sheet(self):
        """Создание главного дашборда с динамическими формулами - ИСПРАВЛЕННАЯ ВЕРСИЯ"""
        ws = self.wb.create_sheet("📊 ГЛАВНЫЙ ДАШБОРД", 0)
        ws.sheet_view.showGridLines = False
        
        # === ЗАГОЛОВОК ===
        ws.merge_cells('A1:Q3')
        title_cell = ws['A1']
        title_cell.value = "🌟 ВАШ ФИНАНСОВЫЙ КОМПАНЬОН ПРОФЕССИОНАЛ"
        title_cell.font = Font(name='Calibri', size=28, bold=True, color='FFFFFFFF')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        title_cell.fill = PatternFill(
            start_color=self.colors['primary']['dark_blue'],
            end_color=self.colors['primary']['dark_blue'],
            fill_type="solid"
        )
        
        # === ДАТА И СОВЕТ ===
        ws['A5'] = f"📅 Отчет создан: {self.today.strftime('%d.%m.%Y %H:%M')}"
        ws['A5'].font = Font(size=11, color=self.colors['primary']['dark_blue'])
        
        ws['A6'] = f"💡 Совет дня: {self.get_tip_of_the_day()}"
        ws['A6'].font = Font(size=11, italic=True, color=self.colors['accent']['info'])
        
        # === КАРТОЧКИ С ДИНАМИЧЕСКИМИ МЕТРИКАМИ ===
        metrics_start = 8
        
        # Подготовим формулы
        income_formula = "=SUM('💰 ДОХОДЫ'!C:C)" if self.incomes else "0"
        
        # Карточка 1: Общий доход
        self.create_metric_card(
            ws, metrics_start, 1,
            "📈 💰 ОБЩИЙ ДОХОД",
            income_formula,
            "в месяц",
            self.colors['accent']['success']
        )
        
        # Карточка 2: Кредитная нагрузка
        if self.credits:
            payment_refs = []
            for credit in self.credits:
                clean_name = str(credit.get('name', '')).replace("'", "").replace('"', '').replace(':', '')
                sheet_name = f"{credit.get('id', 1)}. {clean_name}"[:31]
                payment_refs.append(f"'{sheet_name}'!C23")
            
            if payment_refs:
                credit_load_formula = f"SUM({','.join(payment_refs)})"
            else:
                credit_load_formula = "0"
            load_display = f"ROUND({credit_load_formula}/A{metrics_start+1}*100,1)"
            load_display_formula = f'=TEXT({load_display},"0.0") & "% от дохода"'
        else:
            credit_load_formula = "0"
            load_display_formula = "Нет кредитов"
        
        self.create_metric_card(
            ws, metrics_start + 4, 1,
            "⚖️ 💳 КРЕДИТНАЯ НАГРУЗКА",
            f"={credit_load_formula}",
            load_display_formula if credit_load_formula != "0" else "Нет кредитов",
            self.colors['accent']['warning']
        )
        
        # Карточка 3: Финансовое здоровье
        if self.credits:
            credit = self.credits[0]
            clean_name = str(credit.get('name', '')).replace("'", "").replace('"', '').replace(':', '')
            sheet_name = f"{credit.get('id', 1)}. {clean_name}"[:31]
            
            health_display = f"'{sheet_name}'!B14"
            health_grade = f"IF('{sheet_name}'!B14>=90,'🏆 Отличное',IF('{sheet_name}'!B14>=75,'✅ Хорошее',IF('{sheet_name}'!B14>=60,'⚠️ Требует внимания','🚨 Критическое')))"
        else:
            health_display = "Нет данных"
            health_grade = ""
        
        self.create_metric_card(
            ws, metrics_start, 6,
            "❤️ 🏆 ФИНАНСОВОЕ ЗДОРОВЬЕ",
            f"={health_display}",
            f"={health_grade}" if health_grade else "Нет данных",
            self.colors['gradients']['blue'][0]
        )
        
        # Карточка 4: Выполнено
        if self.credits:
            credit = self.credits[0]
            clean_name = str(credit.get('name', '')).replace("'", "").replace('"', '').replace(':', '')
            sheet_name = f"{credit.get('id', 1)}. {clean_name}"[:31]
            
            progress_percent = f"'{sheet_name}'!B12"
            progress_display = f"'{sheet_name}'!B8 & ' ₽ из ' & '{sheet_name}'!B4 & ' ₽'"
        else:
            progress_percent = "0"
            progress_display = "Нет кредитов"
        
        self.create_metric_card(
            ws, metrics_start + 4, 6,
            "✅ 🎯 ВЫПОЛНЕНО",
            f"={progress_percent}",
            f"={progress_display}" if progress_percent != "0" else "Нет кредитов",
            self.colors['gradients']['green'][0]
        )
        
        # Карточка 5: Свободные средства
        if self.credits and 'payment_refs' in locals() and payment_refs:
            credit_sum = f"SUM({','.join(payment_refs)})"
            expense_sum = "SUM('📊 РАСХОДЫ'!C:C)" if self.expenses else "0"
            free_cash_formula = f"A{metrics_start+1}-{credit_sum}-{expense_sum}"
        else:
            expense_sum = "SUM('📊 РАСХОДЫ'!C:C)" if self.expenses else "0"
            free_cash_formula = f"A{metrics_start+1}-{expense_sum}"
        
        self.create_metric_card(
            ws, metrics_start, 11,
            "💫 ⏳ СВОБОДНЫХ СРЕДСТВ",
            f"={free_cash_formula}",
            "в месяц после всех платежей",
            self.colors['accent']['purple']
        )
        
        # Карточка 6: Цели и накопления
        if self.goals:
            goals_count = f"COUNTA('🎯 ЦЕЛИ'!B:B)-1"
            goals_current = "SUM('🎯 ЦЕЛИ'!D:D)"
            goals_target = "SUM('🎯 ЦЕЛИ'!C:C)"
            goals_percent = f"ROUND({goals_current}/{goals_target}*100,1)"
            goals_display = f"{goals_count} & ' целей'"
            goals_progress = f"{goals_percent} & '% достигнуто'"
        else:
            goals_display = "Нет целей"
            goals_progress = ""
        
        self.create_metric_card(
            ws, metrics_start + 4, 11,
            "🎯 📊 ЦЕЛИ И НАКОПЛЕНИЯ",
            f"={goals_display}",
            f"={goals_progress}" if goals_progress else "Нет целей",
            self.colors['gradients']['purple'][0]
        )
        
        # === ТАБЛИЦА КРЕДИТОВ ===
        if self.credits:
            table_start = metrics_start + 9
            
            ws.merge_cells(f'A{table_start}:Q{table_start}')
            ws[f'A{table_start}'] = "📋 ОБЗОР ВАШИХ КРЕДИТОВ"
            ws[f'A{table_start}'].font = Font(size=16, bold=True, color=self.colors['primary']['dark_blue'])
            ws[f'A{table_start}'].alignment = Alignment(horizontal='center')
            
            # Заголовки таблицы
            headers = ['№', 'Кредит', 'Банк', 'Тип', 'Сумма', 'Ставка', 'Срок', 
                      'Начало', 'Оплачено', 'Остаток', 'Прогресс', 'Платеж', 
                      'Здоровье', 'Статус', 'Рекомендация']
            
            header_row = table_start + 2
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=header_row, column=col, value=header)
                cell.font = Font(bold=True, color='FFFFFFFF')
                cell.fill = PatternFill(start_color=self.colors['primary']['dark_blue'], 
                                       end_color=self.colors['primary']['dark_blue'], 
                                       fill_type='solid')
                cell.alignment = Alignment(horizontal='center')
                cell.border = Border(bottom=Side(style='thin', color='FFFFFFFF'))
            
            # Данные кредитов
            for i, credit in enumerate(self.credits, 1):
                row = header_row + i
                
                clean_credit_name = str(credit.get('name', '')).replace("'", "").replace('"', '').replace(':', '')
                sheet_name = f"{credit.get('id', i)}. {clean_credit_name}"[:31]
                
                # Формулы для ссылок на лист кредита
                data = [
                    i,
                    credit.get('name', ''),
                    credit.get('bank', ''),
                    credit.get('type', ''),
                    f"='{sheet_name}'!$B$4",
                    f"='{sheet_name}'!$D$17",
                    f"='{sheet_name}'!$F$17",
                    f"='{sheet_name}'!$H$17",
                    f"='{sheet_name}'!$B$8",
                    f"='{sheet_name}'!$B$10",
                    f"='{sheet_name}'!$B$12",
                    f"='{sheet_name}'!$C$23",
                    f"='{sheet_name}'!$B$14",
                    f"=IF(M{row}>=90,'🏆 Отличное',IF(M{row}>=75,'✅ Хорошее',IF(M{row}>=60,'⚠️ Требует внимания','🚨 Критическое')))",
                    f"=IF(M{row}>=90,'Продолжайте в том же духе!',IF(M{row}>=75,'Все хорошо!',IF(M{row}>=60,'Требует внимания','Требуются срочные меры!')))"
                ]
                
                for col_idx, value in enumerate(data, 1):
                    cell = ws.cell(row=row, column=col_idx, value=value)
                    
                    if col_idx in [5, 9, 10, 12]:
                        cell.number_format = '#,##0 ₽'
                        cell.alignment = Alignment(horizontal='right')
                    elif col_idx == 6:
                        cell.number_format = '0.0"%"'
                        cell.alignment = Alignment(horizontal='right')
                    elif col_idx == 11:
                        cell.number_format = '0.0%'
                        cell.alignment = Alignment(horizontal='right')
                    elif col_idx == 13:
                        cell.number_format = '0.0'
                        cell.alignment = Alignment(horizontal='right')
        
        # Настройка ширины колонок
        widths = [5, 20, 15, 12, 12, 8, 8, 10, 12, 12, 10, 12, 10, 12, 15]
        for i, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
    
    def create_metric_card(self, ws, row, col, title, value_formula, subtitle_formula, color):
        """Создание карточки с метрикой - ИСПРАВЛЕННАЯ ВЕРСИЯ"""
        # Объединяем ячейки для заголовка
        ws.merge_cells(start_row=row, start_column=col, 
                      end_row=row, end_column=col+3)
        
        title_cell = ws.cell(row=row, column=col)
        title_cell.value = title
        title_cell.font = Font(name='Calibri', size=11, bold=True, color='FFFFFFFF')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        title_cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        
        # Объединяем ячейки для значения
        ws.merge_cells(start_row=row+1, start_column=col, 
                      end_row=row+1, end_column=col+3)
        
        value_cell = ws.cell(row=row+1, column=col)
        if value_formula and value_formula.startswith('='):
            value_cell.value = value_formula
        elif value_formula:
            value_cell.value = f"={value_formula}"
        
        value_cell.font = Font(name='Calibri', size=14, bold=True, color='FFFFFFFF')
        value_cell.alignment = Alignment(horizontal='center', vertical='center')
        value_cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        
        # Объединяем ячейки для подзаголовка
        ws.merge_cells(start_row=row+2, start_column=col, 
                      end_row=row+2, end_column=col+3)
        
        subtitle_cell = ws.cell(row=row+2, column=col)
        if subtitle_formula and subtitle_formula.startswith('='):
            subtitle_cell.value = subtitle_formula
        elif subtitle_formula:
            subtitle_cell.value = subtitle_formula
        
        subtitle_cell.font = Font(name='Calibri', size=10, color='FFFFFFFF')
        subtitle_cell.alignment = Alignment(horizontal='center', vertical='center')
        subtitle_cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        
        # Рамка для всей карточки
        for r in range(row, row+3):
            for c in range(col, col+4):
                ws.cell(row=r, column=c).border = Border(
                    left=Side(style='thin', color='FFFFFFFF'),
                    right=Side(style='thin', color='FFFFFFFF'),
                    top=Side(style='thin', color='FFFFFFFF'),
                    bottom=Side(style='thin', color='FFFFFFFF')
                )
    
    def create_credit_sheets(self):
        """Создание отдельных листов для каждого кредита"""
        for credit in self.credits:
            self.create_individual_credit_sheet(credit)
    
    def create_individual_credit_sheet(self, credit):
        """Создание детального листа для кредита - ИСПРАВЛЕННАЯ ВЕРСИЯ"""
        clean_name = str(credit.get('name', '')).replace('/', '').replace('\\', '').replace('?', '').replace('*', '').replace(':', '').replace('[', '').replace(']', '')
        sheet_name = f"{credit.get('id', 1)}. {clean_name}"[:31]
        
        ws = self.wb.create_sheet(sheet_name)
        ws.sheet_view.showGridLines = False
        
        # === ЗАГОЛОВОК ===
        ws.merge_cells('A1:G2')
        title = ws['A1']
        title.value = f"💎 {credit.get('name', '').upper()}"
        title.font = Font(name='Calibri', size=18, bold=True, color='FFFFFFFF')
        title.alignment = Alignment(horizontal='center', vertical='center')
        title.fill = PatternFill(
            start_color=self.colors['primary']['dark_blue'],
            end_color=self.colors['primary']['dark_blue'],
            fill_type="solid"
        )
        
        # === КЛЮЧЕВЫЕ ПОКАЗАТЕЛИ ===
        ws['A4'] = "Общая сумма:"
        ws['A4'].font = Font(bold=True)
        ws['B4'] = credit.get('amount', 0)
        ws['B4'].number_format = '#,##0 ₽'
        ws['B4'].font = Font(bold=True, color=self.colors['primary']['dark_blue'])
        
        ws['A8'] = "Уже оплачено:"
        ws['A8'].font = Font(bold=True)
        ws['B8'] = credit.get('paid', 0)
        ws['B8'].number_format = '#,##0 ₽'
        ws['B8'].font = Font(bold=True, color=self.colors['accent']['success'])
        
        ws['A10'] = "Остаток:"
        ws['A10'].font = Font(bold=True)
        ws['B10'] = f"=B4-B8"
        ws['B10'].number_format = '#,##0 ₽'
        ws['B10'].font = Font(bold=True, color=self.colors['accent']['warning'])
        
        ws['A12'] = "Прогресс:"
        ws['A12'].font = Font(bold=True)
        ws['B12'] = "=IF(B4=0,0,B8/B4)"
        ws['B12'].number_format = '0.0%'
        ws['B12'].font = Font(bold=True, color=self.colors['accent']['info'])
        
        # Рассчитываем здоровье
        health_score = self.calculate_financial_health_score(credit)
        ws['A14'] = "Здоровье (%):"
        ws['A14'].font = Font(bold=True)
        ws['B14'] = health_score['total']
        ws['B14'].number_format = '0.0'
        ws['B14'].font = Font(bold=True, color=self.colors['accent']['purple'])
        
        # === ИНФОКАРТОЧКИ ===
        info_start = 17
        info_cards = [
            {
                'title': '🏦 Банк',
                'value': credit.get('bank', 'Не указан'),
                'color': self.colors['accent']['info'],
                'col': 1
            },
            {
                'title': '📈 Ставка',
                'value': f"{credit.get('rate', 0)*100:.1f}%",
                'color': self.colors['accent']['warning'],
                'col': 3
            },
            {
                'title': '📅 Срок',
                'value': f"{credit.get('months', 0)} мес",
                'color': self.colors['accent']['purple'],
                'col': 5
            },
            {
                'title': '🗓️ Начало',
                'value': credit.get('start_date', '') if isinstance(credit.get('start_date', ''), str) else credit.get('start_date', '').strftime('%d.%m.%Y'),
                'color': self.colors['accent']['success'],
                'col': 7
            }
        ]
        
        for card in info_cards:
            ws.merge_cells(start_row=info_start, start_column=card['col'],
                          end_row=info_start+2, end_column=card['col'])
            
            cell = ws.cell(row=info_start, column=card['col'])
            cell.value = f"{card['title']}\n{card['value']}"
            cell.font = Font(bold=True, color='FFFFFFFF', size=10)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.fill = PatternFill(start_color=card['color'], 
                                   end_color=card['color'], 
                                   fill_type='solid')
        
        # === ТАБЛИЦА ПЛАТЕЖЕЙ ===
        table_start = info_start + 5
        
        ws.merge_cells(f'A{table_start}:G{table_start}')
        ws[f'A{table_start}'] = "📋 ГРАФИК ПЛАТЕЖЕЙ"
        ws[f'A{table_start}'].font = Font(bold=True, color=self.colors['primary']['dark_blue'])
        
        headers = ['Месяц', 'Дата', 'Платеж', 'Оплачено', 'Остаток', 'Статус', '🎯']
        header_row = table_start + 2
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFFFF')
            cell.fill = PatternFill(start_color=self.colors['primary']['dark_blue'], 
                                   end_color=self.colors['primary']['dark_blue'], 
                                   fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
            cell.border = Border(bottom=Side(style='thin', color='FFFFFFFF'))
        
        # Заполняем таблицу - ИСПРАВЛЕННЫЙ РАСЧЕТ ДАТ
        monthly_payment = credit.get('monthly_payment', self.calculate_monthly_payment(credit))
        max_months_to_show = min(credit.get('months', 12), 12)
        
        start_date = credit.get('start_date', self.today)
        if isinstance(start_date, str):
            try:
                start_date = datetime.strptime(start_date, '%d.%m.%Y')
            except:
                start_date = self.today
        
        for month in range(1, max_months_to_show + 1):
            row = header_row + month
            
            # Дата платежа (правильный расчет)
            payment_date = start_date + relativedelta(months=month-1)
            
            # Месяц
            ws.cell(row=row, column=1, value=month).alignment = Alignment(horizontal='center')
            
            # Дата
            date_cell = ws.cell(row=row, column=2, value=payment_date)
            date_cell.number_format = 'DD.MM.YYYY'
            
            # Платеж
            ws.cell(row=row, column=3, value=monthly_payment)
            ws.cell(row=row, column=3).number_format = '#,##0 ₽'
            
            # Оплачено (для ввода пользователем)
            paid_cell = ws.cell(row=row, column=4, value='')
            paid_cell.number_format = '#,##0 ₽'
            
            # Остаток
            if month == 1:
                formula = f"=$B$4-D{row}"
            else:
                formula = f"=E{row-1}-D{row}"
            balance_cell = ws.cell(row=row, column=5, value=formula)
            balance_cell.number_format = '#,##0 ₽'
            
            # Статус
            status_formula = f'=IF(D{row}>=C{row},"✅",IF(D{row}>0,"⚠️","⏳"))'
            ws.cell(row=row, column=6, value=status_formula)
            ws.cell(row=row, column=6).alignment = Alignment(horizontal='center')
            
            # Мотивационная колонка
            if month <= self.calculate_months_passed(credit):
                ws.cell(row=row, column=7, value="🎉")
            elif month == self.calculate_months_passed(credit) + 1:
                ws.cell(row=row, column=7, value="📅")
            else:
                ws.cell(row=row, column=7, value="⏳")
            ws.cell(row=row, column=7).alignment = Alignment(horizontal='center')
        
        # Итоги
        total_row = header_row + max_months_to_show + 2
        ws[f'A{total_row}'] = "Итого выплат:"
        ws[f'B{total_row}'] = f"=SUM(D{header_row+1}:D{header_row+max_months_to_show})"
        ws[f'B{total_row}'].number_format = '#,##0 ₽'
        ws[f'B{total_row}'].font = Font(bold=True, color=self.colors['accent']['success'])
        
        ws[f'A{total_row+1}'] = "Переплата:"
        ws[f'B{total_row+1}'] = f"=MAX(0, B{total_row}-$B$4)"
        ws[f'B{total_row+1}'].number_format = '#,##0 ₽'
        
        # ДОБАВЛЕНО: Полная стоимость кредита
        ws[f'A{total_row+2}'] = "Полная стоимость:"
        ws[f'B{total_row+2}'] = f"=B{total_row}"
        ws[f'B{total_row+2}'].number_format = '#,##0 ₽'
        
        ws[f'A{total_row+3}'] = "Общая переплата:"
        ws[f'B{total_row+3}'] = f"=B{total_row+1}"
        ws[f'B{total_row+3}'].number_format = '#,##0 ₽'
        ws[f'B{total_row+3}'].font = Font(
            color=self.colors['accent']['danger'] if monthly_payment * credit.get('months', 0) > credit.get('amount', 0) 
            else self.colors['accent']['success']
        )
        
        # Настройка ширины
        widths = [8, 12, 12, 12, 12, 8, 6]
        for i, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
    
    def create_summary_sheet(self):
        """Лист с краткой сводкой"""
        ws = self.wb.create_sheet("📋 СВОДКА")
        ws.sheet_view.showGridLines = False
        
        # Заголовок
        ws.merge_cells('A1:D1')
        ws['A1'] = "🌟 БЫСТРАЯ СВОДКА ПО ФИНАНСАМ"
        ws['A1'].font = Font(size=20, bold=True, color=self.colors['primary']['dark_blue'])
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Основные показатели
        ws['A3'] = "📊 КЛЮЧЕВЫЕ ПОКАЗАТЕЛИ"
        ws['A3'].font = Font(size=14, bold=True, color=self.colors['accent']['info'])
        
        total_income = sum(i.get('amount', 0) for i in self.incomes) if self.incomes else 0
        total_monthly_payments = sum(c.get('monthly_payment', 0) for c in self.credits) if self.credits else 0
        total_credits = sum(c.get('amount', 0) for c in self.credits) if self.credits else 0
        total_paid = sum(c.get('paid', 0) for c in self.credits) if self.credits else 0
        
        metrics = [
            ["💰 Общий доход", f"{total_income:,.0f} ₽/мес" if self.incomes else "Не указан"],
            ["💳 Кредитная нагрузка", f"{total_monthly_payments:,.0f} ₽/мес" if self.credits else "Нет кредитов"],
            ["📈 Прогресс по кредитам", f"{(total_paid/total_credits*100 if total_credits>0 else 0):.1f}%" if self.credits else "Нет кредитов"],
            ["🎯 Активных целей", f"{len(self.goals)}" if self.goals else "Нет целей"],
            ["📅 Дата следующего платежа", self.get_next_payment_date()],
            ["💫 Свободные средства", f"{(total_income - total_monthly_payments):,.0f} ₽/мес" if self.incomes else "Не рассчитано"]
        ]
        
        for i, (label, value) in enumerate(metrics, 5):
            ws[f'A{i}'] = label
            ws[f'A{i}'].font = Font(bold=True)
            ws[f'C{i}'] = value
            ws[f'C{i}'].font = Font(size=12)
    
    def get_next_payment_date(self):
        """Дата следующего платежа"""
        if not self.credits:
            return "Нет активных кредитов"
        
        next_dates = []
        for credit in self.credits:
            months_passed = self.calculate_months_passed(credit)
            if months_passed < credit.get('months', 0):
                start_date = credit.get('start_date', self.today)
                if isinstance(start_date, str):
                    try:
                        start_date = datetime.strptime(start_date, '%d.%m.%Y')
                    except:
                        start_date = self.today
                next_payment = start_date + relativedelta(months=months_passed)
                next_dates.append(next_payment)
        
        if next_dates:
            return min(next_dates).strftime('%d.%m.%Y')
        return "Все кредиты погашены"
    
    def create_income_sheet(self):
        """Лист доходов"""
        if not self.incomes:
            return
        
        ws = self.wb.create_sheet("💰 ДОХОДЫ")
        ws.sheet_view.showGridLines = False
        
        ws.merge_cells('A1:E1')
        ws['A1'] = "💰 ВАШИ ДОХОДЫ"
        ws['A1'].font = Font(size=18, bold=True, color=self.colors['primary']['dark_blue'])
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Заголовки таблицы
        headers = ['№', 'Источник', 'Сумма/мес', 'Категория', 'Стабильность']
        start_row = 3
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFFFF')
            cell.fill = PatternFill(start_color=self.colors['primary']['dark_blue'], 
                                   end_color=self.colors['primary']['dark_blue'], 
                                   fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        # Данные доходов
        for i, income in enumerate(self.incomes, 1):
            row = start_row + i
            ws.cell(row=row, column=1, value=i)
            ws.cell(row=row, column=2, value=income.get('name', ''))
            ws.cell(row=row, column=3, value=income.get('amount', 0)).number_format = '#,##0 ₽'
            ws.cell(row=row, column=4, value=income.get('category', ''))
            ws.cell(row=row, column=5, value=income.get('stability', ''))
        
        # Итог
        total_row = start_row + len(self.incomes) + 2
        ws.cell(row=total_row, column=2, value="ИТОГО:").font = Font(bold=True)
        ws.cell(row=total_row, column=3, value=sum(i.get('amount', 0) for i in self.incomes)).number_format = '#,##0 ₽'
        ws.cell(row=total_row, column=3).font = Font(bold=True, color=self.colors['accent']['success'])
        
        # Настройка ширины
        widths = [5, 25, 15, 15, 15]
        for i, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
    
    def create_expense_sheet(self):
        """Лист расходов"""
        if not self.expenses:
            return
        
        ws = self.wb.create_sheet("📊 РАСХОДЫ")
        ws.sheet_view.showGridLines = False
        
        ws.merge_cells('A1:D1')
        ws['A1'] = "📊 ВАШИ РАСХОДЫ"
        ws['A1'].font = Font(size=18, bold=True, color=self.colors['primary']['dark_blue'])
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Заголовки таблицы
        headers = ['№', 'Статья расходов', 'Сумма/мес', 'Категория']
        start_row = 3
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFFFF')
            cell.fill = PatternFill(start_color=self.colors['primary']['dark_blue'], 
                                   end_color=self.colors['primary']['dark_blue'], 
                                   fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        # Данные расходов
        for i, expense in enumerate(self.expenses, 1):
            row = start_row + i
            ws.cell(row=row, column=1, value=i)
            ws.cell(row=row, column=2, value=expense.get('name', ''))
            ws.cell(row=row, column=3, value=expense.get('amount', 0)).number_format = '#,##0 ₽'
            ws.cell(row=row, column=4, value=expense.get('category', ''))
        
        # Итог
        total_row = start_row + len(self.expenses) + 2
        ws.cell(row=total_row, column=2, value="ИТОГО:").font = Font(bold=True)
        ws.cell(row=total_row, column=3, value=sum(e.get('amount', 0) for e in self.expenses)).number_format = '#,##0 ₽'
        ws.cell(row=total_row, column=3).font = Font(bold=True, color=self.colors['accent']['warning'])
        
        # Настройка ширины
        widths = [5, 30, 15, 15]
        for i, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
    
    def create_goals_sheet(self):
        """Лист финансовых целей"""
        if not self.goals:
            return
        
        ws = self.wb.create_sheet("🎯 ЦЕЛИ")
        ws.sheet_view.showGridLines = False
        
        ws.merge_cells('A1:H1')
        ws['A1'] = "🎯 ВАШИ ФИНАНСОВЫЕ ЦЕЛИ"
        ws['A1'].font = Font(size=18, bold=True, color=self.colors['primary']['dark_blue'])
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Заголовки таблицы
        headers = ['№', 'Цель', 'Целевая сумма', 'Накоплено', 'Осталось', 'Срок', 'Прогресс', 'Приоритет']
        start_row = 3
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFFFF')
            cell.fill = PatternFill(start_color=self.colors['primary']['dark_blue'], 
                                   end_color=self.colors['primary']['dark_blue'], 
                                   fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        # Данные целей
        for i, goal in enumerate(self.goals, 1):
            row = start_row + i
            remaining = goal.get('target', 0) - goal.get('current', 0)
            progress = (goal.get('current', 0) / goal.get('target', 1)) if goal.get('target', 0) > 0 else 0
            
            ws.cell(row=row, column=1, value=i)
            ws.cell(row=row, column=2, value=goal.get('name', ''))
            ws.cell(row=row, column=3, value=goal.get('target', 0)).number_format = '#,##0 ₽'
            ws.cell(row=row, column=4, value=goal.get('current', 0)).number_format = '#,##0 ₽'
            ws.cell(row=row, column=5, value=remaining).number_format = '#,##0 ₽'
            
            deadline = goal.get('deadline', '')
            if isinstance(deadline, str):
                deadline_str = deadline
            else:
                deadline_str = deadline.strftime('%d.%m.%Y')
            ws.cell(row=row, column=6, value=deadline_str)
            
            ws.cell(row=row, column=7, value=progress).number_format = '0.0%'
            ws.cell(row=row, column=8, value=goal.get('priority', ''))
            
            # Условное форматирование для прогресса
            if progress >= 1:
                ws.cell(row=row, column=7).font = Font(color=self.colors['accent']['success'], bold=True)
            elif progress >= 0.5:
                ws.cell(row=row, column=7).font = Font(color=self.colors['accent']['warning'], bold=True)
        
        # Итог
        total_row = start_row + len(self.goals) + 2
        ws.cell(row=total_row, column=2, value="ИТОГО:").font = Font(bold=True)
        ws.cell(row=total_row, column=3, value=sum(g.get('target', 0) for g in self.goals)).number_format = '#,##0 ₽'
        ws.cell(row=total_row, column=4, value=sum(g.get('current', 0) for g in self.goals)).number_format = '#,##0 ₽'
        
        # Настройка ширины
        widths = [5, 25, 15, 15, 15, 12, 10, 12]
        for i, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
    
    def create_analytics_sheet(self):
        """Лист аналитики - ДОБАВЛЕН АНАЛИЗ РИСКОВ"""
        ws = self.wb.create_sheet("📈 АНАЛИТИКА")
        ws.sheet_view.showGridLines = False
        
        ws.merge_cells('A1:F1')
        ws['A1'] = "📈 ФИНАНСОВАЯ АНАЛИТИКА"
        ws['A1'].font = Font(size=18, bold=True, color=self.colors['primary']['dark_blue'])
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Анализ финансового состояния
        ws['A3'] = "📊 АНАЛИЗ ВАШЕГО ФИНАНСОВОГО СОСТОЯНИЯ"
        ws['A3'].font = Font(size=14, bold=True, color=self.colors['accent']['info'])
        
        total_income = sum(i.get('amount', 0) for i in self.incomes) if self.incomes else 0
        total_expenses = sum(e.get('amount', 0) for e in self.expenses) if self.expenses else 0
        total_credits_monthly = sum(c.get('monthly_payment', 0) for c in self.credits) if self.credits else 0
        
        analysis = [
            ["💰 Доходы:", f"{total_income:,.0f} ₽/мес" if self.incomes else "Не указано"],
            ["💸 Расходы:", f"{total_expenses:,.0f} ₽/мес" if self.expenses else "Не указано"],
            ["💳 Кредиты:", f"{total_credits_monthly:,.0f} ₽/мес" if self.credits else "Нет кредитов"],
            ["💫 Свободные средства:", f"{(total_income - total_expenses - total_credits_monthly):,.0f} ₽/мес" if total_income > 0 else "Не рассчитано"],
            ["⚖️ Нагрузка на доход:", f"{(total_credits_monthly/total_income*100 if total_income>0 else 0):.1f}%" if self.credits and total_income > 0 else "Не рассчитано"]
        ]
        
        for i, (label, value) in enumerate(analysis, 5):
            ws[f'A{i}'] = label
            ws[f'A{i}'].font = Font(bold=True)
            ws[f'C{i}'] = value
            ws[f'C{i}'].font = Font(size=12)
        
        # ДОБАВЛЕНО: Анализ рисков
        risk_start = len(analysis) + 7
        ws[f'A{risk_start}'] = "📊 АНАЛИЗ РИСКОВ"
        ws[f'A{risk_start}'].font = Font(size=14, bold=True, color=self.colors['accent']['danger'])
        
        # Рассчитываем риски
        credit_share = (total_credits_monthly / total_income) * 100 if total_income > 0 else 0
        free_cash = total_income - total_credits_monthly - total_expenses
        free_cash_ratio = (free_cash / total_income) * 100 if total_income > 0 else 0
        safety_buffer = free_cash / total_credits_monthly if total_credits_monthly > 0 else 0
        
        risks = [
            ["Кредитная нагрузка:", f"{credit_share:.1f}%", 
             "✅ Низкая" if credit_share < 20 else "⚠️ Средняя" if credit_share < 40 else "🚨 Высокая"],
            ["Свободные средства:", f"{free_cash_ratio:.1f}% от дохода",
             "✅ Хорошо" if free_cash_ratio > 20 else "⚠️ Нормально" if free_cash_ratio > 10 else "🚨 Критично"],
            ["Запас прочности:", f"{safety_buffer:.1f} мес" if total_credits_monthly > 0 else "∞",
             "✅ Отлично" if safety_buffer > 3 else "⚠️ Нормально" if safety_buffer > 1 else "🚨 Риск"]
        ]
        
        for i, (label, value, status) in enumerate(risks, risk_start + 2):
            ws[f'A{i}'] = label
            ws[f'A{i}'].font = Font(bold=True)
            ws[f'C{i}'] = value
            ws[f'E{i}'] = status
            
            # Цвет статуса
            if "✅" in status:
                ws[f'E{i}'].font = Font(color=self.colors['accent']['success'], bold=True)
            elif "⚠️" in status:
                ws[f'E{i}'].font = Font(color=self.colors['accent']['warning'], bold=True)
            else:
                ws[f'E{i}'].font = Font(color=self.colors['accent']['danger'], bold=True)
        
        # Настройка ширины
        widths = [25, 5, 20, 5, 20]
        for i, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
    
    def create_calendar_sheet(self):
        """Лист календаря платежей - ИСПРАВЛЕННАЯ ВЕРСИЯ"""
        ws = self.wb.create_sheet("📅 КАЛЕНДАРЬ")
        ws.sheet_view.showGridLines = False
        
        ws.merge_cells('A1:F1')
        ws['A1'] = "📅 КАЛЕНДАРЬ ПЛАТЕЖЕЙ НА 90 ДНЕЙ"
        ws['A1'].font = Font(size=18, bold=True, color=self.colors['primary']['dark_blue'])
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Заголовки
        headers = ['Дата', 'День недели', 'Кредит', 'Сумма платежа', 'Статус', 'До платежа']
        start_row = 3
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFFFF')
            cell.fill = PatternFill(start_color=self.colors['primary']['dark_blue'], 
                                   end_color=self.colors['primary']['dark_blue'], 
                                   fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        # Заполняем календарь на 90 дней
        row = start_row + 1
        days_ru = ['Пн', 'Вт', 'Ср', 'Чт', 'Пт', 'Сб', 'Вс']
        
        for day in range(90):
            current_date = self.today + timedelta(days=day)
            
            # Дата (ТОЛЬКО дата, без времени)
            date_cell = ws.cell(row=row, column=1, value=current_date.date())
            date_cell.number_format = 'DD.MM.YYYY'
            
            # День недели
            day_of_week = days_ru[current_date.weekday()]
            ws.cell(row=row, column=2, value=day_of_week)
            
            # Проверяем платежи
            payments_today = []
            for credit in self.credits:
                months_passed = self.calculate_months_passed(credit)
                if months_passed < credit.get('months', 0):
                    start_date = credit.get('start_date', self.today)
                    if isinstance(start_date, str):
                        try:
                            start_date = datetime.strptime(start_date, '%d.%m.%Y')
                        except:
                            start_date = self.today
                    next_payment_date = start_date + relativedelta(months=months_passed)
                    if next_payment_date.date() == current_date.date():
                        payments_today.append({
                            'name': credit.get('name', ''),
                            'amount': credit.get('monthly_payment', 0)
                        })
            
            if payments_today:
                for payment in payments_today:
                    ws.cell(row=row, column=3, value=payment['name'][:15])
                    ws.cell(row=row, column=4, value=payment['amount']).number_format = '#,##0 ₽'
                    
                    if current_date.date() < self.today.date():
                        ws.cell(row=row, column=5, value="✅ Прошлый").font = Font(color=self.colors['accent']['success'])
                    elif current_date.date() == self.today.date():
                        ws.cell(row=row, column=5, value="📅 Сегодня").font = Font(color=self.colors['accent']['warning'], bold=True)
                    else:
                        days_left = (current_date - self.today).days
                        ws.cell(row=row, column=5, value=f"⏳ Через {days_left} дн.")
                    
                    row += 1
            else:
                ws.cell(row=row, column=3, value="Нет платежей")
                ws.cell(row=row, column=5, value="📅 Свободный день")
                row += 1
        
        # Настройка ширины
        widths = [12, 12, 20, 15, 15, 12]
        for i, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
    
    def create_instructions_sheet(self):
        """Лист с инструкциями"""
        ws = self.wb.create_sheet("📚 ИНСТРУКЦИЯ")
        ws.sheet_view.showGridLines = False
        
        ws.merge_cells('A1:C1')
        ws['A1'] = "🎯 КАК ПОЛЬЗОВАТЬСЯ ФИНАНСОВЫМ КОМПАНЬОНОМ"
        ws['A1'].font = Font(size=18, bold=True, color=self.colors['primary']['dark_blue'])
        ws['A1'].alignment = Alignment(horizontal='center')
        
        instructions = [
            ["", "", ""],
            ["💡 ОСНОВНЫЕ ПРИНЦИПЫ:", "", ""],
            ["", "• Регулярность", "Обновляйте данные после каждого платежа"],
            ["", "• Честность", "Вносите реальные суммы доходов и расходов"],
            ["", "• Дисциплина", "Следуйте своему финансовому плану"],
            ["", "• Анализ", "Регулярно просматривайте аналитику"],
            ["", "", ""],
            ["🔄 КАК ОБНОВЛЯТЬ ДАННЫЕ:", "", ""],
            ["", "1. Откройте лист нужного кредита", ""],
            ["", "2. В столбце 'Оплачено' введите сумму платежа", ""],
            ["", "3. Все остальные ячейки обновятся автоматически", ""],
            ["", "4. Вернитесь на Дашборд для просмотра статистики", ""],
            ["", "", ""],
            ["🎨 ВИЗУАЛЬНЫЕ ПОДСКАЗКИ:", "", ""],
            ["", "✅ Зеленый", "Платеж выполнен полностью"],
            ["", "⚠️ Желтый", "Платеж выполнен частично"],
            ["", "⏳ Серый", "Платеж ожидается"],
            ["", "📊 Цветная шкала", "Визуализация прогресса"],
            ["", "", ""],
            ["🚀 ДЛЯ ЛУЧШИХ РЕЗУЛЬТАТОВ:", "", ""],
            ["", "• Установите напоминания о платежах", ""],
            ["", "• Планируйте платежи на одну дату", ""],
            ["", "• Создайте финансовую подушку", ""],
            ["", "• Отмечайте достижения", ""],
        ]
        
        start_row = 3
        for i, (col1, col2, col3) in enumerate(instructions):
            row = start_row + i
            ws[f'A{row}'] = col1
            ws[f'B{row}'] = col2
            ws[f'C{row}'] = col3
            
            if col1 and "💡" in col1:
                ws[f'A{row}'].font = Font(bold=True, color=self.colors['accent']['info'])
            if col1 and "🔄" in col1:
                ws[f'A{row}'].font = Font(bold=True, color=self.colors['accent']['warning'])
            if col1 and "🎨" in col1:
                ws[f'A{row}'].font = Font(bold=True, color=self.colors['accent']['purple'])
            if col1 and "🚀" in col1:
                ws[f'A{row}'].font = Font(bold=True, color=self.colors['accent']['success'])
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 40

    def complete_visual_upgrade(self):
        """Полное визуальное обновление файла"""
        print("   🎨 Запуск визуального обновления...")
        
        try:
            self.enhance_visual_design()
            self.adjust_column_widths()
            print("   ✅ Визуальное обновление завершено!")
            
        except Exception as e:
            print(f"   ⚠️ Ошибка при визуальном обновлении: {e}")
    
    def enhance_visual_design(self):
        """Улучшение визуального дизайна всех листов"""
        try:
            for sheet_name in self.wb.sheetnames:
                ws = self.wb[sheet_name]
                
                for row in ws.iter_rows(min_row=1, max_row=10):
                    for cell in row:
                        if cell.value and isinstance(cell.value, str):
                            if any(emoji in cell.value for emoji in ['💰', '💳', '📊', '🎯', '📈', '📅', '📚', '🌟']):
                                if '🌟' in cell.value:
                                    cell.font = Font(size=cell.font.size or 14, bold=True, 
                                                   color=self.colors['primary']['dark_blue'])
                                else:
                                    cell.font = Font(bold=True, color=self.colors['primary']['dark_blue'])
                
                for row in range(1, ws.max_row + 1):
                    if row % 2 == 0:
                        for col in range(1, ws.max_column + 1):
                            cell = ws.cell(row=row, column=col)
                            if cell.value and cell.fill.start_color.index in ['00000000', None]:
                                cell.fill = PatternFill(start_color=self.colors['background']['gray_light'], 
                                                      end_color=self.colors['background']['gray_light'], 
                                                      fill_type="solid")
        except:
            pass
    
    def adjust_column_widths(self):
        """Автоматическая настройка ширины колонок"""
        try:
            for ws in self.wb.worksheets:
                for column_cells in ws.columns:
                    if column_cells:
                        max_length = 0
                        column = column_cells[0].column
                        
                        for cell in column_cells:
                            try:
                                cell_length = len(str(cell.value or ""))
                                if cell_length > max_length:
                                    max_length = cell_length
                            except:
                                pass
                        
                        adjusted_width = min(max_length + 2, 50)
                        ws.column_dimensions[get_column_letter(column)].width = adjusted_width
        except:
            pass


class FinancialCompanionGUI:
    """Графический интерфейс для Финансового Компаньона"""
    
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("ФИНАНСОВЫЙ КОМПАНЬОН ПРОФЕССИОНАЛ 2026")
        self.root.geometry("1000x700")
        self.root.configure(bg='#F5F7FA')
        
        self.companion = UltimateFinancialCompanion()
        
        self.colors = {
            'primary': '#0F2B46',
            'secondary': '#1E5A8A',
            'success': '#4CAF50',
            'warning': '#FF9800',
            'danger': '#F44336',
            'light': '#F5F7FA',
            'white': '#FFFFFF'
        }
        
        self.incomes = []
        self.credits = []
        self.expenses = []
        self.goals = []
        
        # ИСПРАВЛЕНИЕ: Добавляем атрибут today
        self.today = datetime.now()
        
        self.setup_ui()
        
    def setup_ui(self):
        """Настройка пользовательского интерфейса"""
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill='both', expand=True, padx=20, pady=20)
        
        self.setup_welcome_tab()
        self.setup_income_tab()
        self.setup_credit_tab()
        self.setup_expense_tab()
        self.setup_goal_tab()
        self.setup_summary_tab()
        
        self.setup_footer()
        
    def setup_welcome_tab(self):
        """Вкладка приветствия"""
        welcome_frame = tk.Frame(self.notebook, bg=self.colors['white'])
        self.notebook.add(welcome_frame, text='Добро пожаловать')
        
        title_label = tk.Label(
            welcome_frame,
            text="ФИНАНСОВЫЙ КОМПАНЬОН ПРОФЕССИОНАЛ 2026",
            font=('Arial', 24, 'bold'),
            fg=self.colors['primary'],
            bg=self.colors['white']
        )
        title_label.pack(pady=40)
        
        tip_frame = tk.Frame(welcome_frame, bg='#E3F2FD', relief='groove', borderwidth=2)
        tip_frame.pack(pady=20, padx=50, fill='x')
        
        tip_label = tk.Label(
            tip_frame,
            text="СОВЕТ ДНЯ\n\n" + self.companion.get_tip_of_the_day(),
            font=('Arial', 12),
            fg=self.colors['primary'],
            bg='#E3F2FD',
            wraplength=600,
            justify='left'
        )
        tip_label.pack(pady=20, padx=20)
        
        instruction_text = """КАК ПОЛЬЗОВАТЬСЯ
        
        1. Заполните данные о доходах во вкладке Доходы
        2. Добавьте информацию о кредитах во вкладке Кредиты
        3. Укажите расходы во вкладке Расходы
        4. Поставьте финансовые цели во вкладке Цели
        5. Проверьте сводку во вкладке Итоги
        6. Нажмите Создать отчет для генерации Excel-файла
        
        Все показатели динамические и обновляются автоматически!"""
        
        instruction_frame = tk.Frame(welcome_frame, bg='#FFF3E0', relief='groove', borderwidth=2)
        instruction_frame.pack(pady=20, padx=50, fill='x')
        
        instruction_label = tk.Label(
            instruction_frame,
            text=instruction_text,
            font=('Arial', 11),
            fg=self.colors['primary'],
            bg='#FFF3E0',
            justify='left'
        )
        instruction_label.pack(pady=20, padx=20)
        
        stats_frame = tk.Frame(welcome_frame, bg=self.colors['white'])
        stats_frame.pack(pady=30)
        
        stats_label = tk.Label(
            stats_frame,
            text="ЦЕЛЬ ПРОГРАММЫ: Помочь вам достичь финансовой свободы!",
            font=('Arial', 14, 'bold'),
            fg=self.colors['success'],
            bg=self.colors['white']
        )
        stats_label.pack()
        
    def setup_income_tab(self):
        """Вкладка доходов"""
        income_frame = tk.Frame(self.notebook, bg=self.colors['white'])
        self.notebook.add(income_frame, text='Доходы')
        
        tk.Label(
            income_frame,
            text="ВАШИ ДОХОДЫ",
            font=('Arial', 20, 'bold'),
            fg=self.colors['primary'],
            bg=self.colors['white']
        ).pack(pady=20)
        
        form_frame = tk.Frame(income_frame, bg='#E8F5E9', relief='ridge', borderwidth=2)
        form_frame.pack(pady=10, padx=30, fill='x')
        
        tk.Label(
            form_frame,
            text="Добавить новый доход",
            font=('Arial', 12, 'bold'),
            fg=self.colors['success'],
            bg='#E8F5E9'
        ).pack(pady=10)
        
        fields_frame = tk.Frame(form_frame, bg='#E8F5E9')
        fields_frame.pack(pady=10, padx=20)
        
        tk.Label(fields_frame, text="Название источника", bg='#E8F5E9').grid(row=0, column=0, sticky='w', pady=5)
        self.income_name = tk.Entry(fields_frame, width=30)
        self.income_name.grid(row=0, column=1, pady=5, padx=10)
        
        tk.Label(fields_frame, text="Сумма в месяц (руб)", bg='#E8F5E9').grid(row=1, column=0, sticky='w', pady=5)
        self.income_amount = tk.Entry(fields_frame, width=30)
        self.income_amount.grid(row=1, column=1, pady=5, padx=10)
        
        tk.Label(fields_frame, text="Категория", bg='#E8F5E9').grid(row=2, column=0, sticky='w', pady=5)
        self.income_category = ttk.Combobox(fields_frame, width=28, values=["Основной", "Дополнительный", "Пассивный"])
        self.income_category.grid(row=2, column=1, pady=5, padx=10)
        self.income_category.set("Основной")
        
        tk.Label(fields_frame, text="Стабильность", bg='#E8F5E9').grid(row=3, column=0, sticky='w', pady=5)
        self.income_stability = ttk.Combobox(fields_frame, width=28, values=["Высокая", "Средняя", "Низкая"])
        self.income_stability.grid(row=3, column=1, pady=5, padx=10)
        self.income_stability.set("Высокая")
        
        add_button = tk.Button(
            form_frame,
            text="Добавить доход",
            command=self.add_income,
            bg=self.colors['success'],
            fg='white',
            font=('Arial', 10, 'bold'),
            padx=20,
            pady=10
        )
        add_button.pack(pady=20)
        
        table_frame = tk.Frame(income_frame, bg=self.colors['white'])
        table_frame.pack(pady=20, padx=30, fill='both', expand=True)
        
        tk.Label(
            table_frame,
            text="СПИСОК ДОХОДОВ",
            font=('Arial', 14, 'bold'),
            fg=self.colors['primary'],
            bg=self.colors['white']
        ).pack(anchor='w', pady=10)
        
        columns = ('#', 'Источник', 'Сумма/мес', 'Категория', 'Стабильность')
        self.income_tree = ttk.Treeview(table_frame, columns=columns, show='headings', height=8)
        
        for col in columns:
            self.income_tree.heading(col, text=col)
            self.income_tree.column(col, width=120)
        
        scrollbar = ttk.Scrollbar(table_frame, orient='vertical', command=self.income_tree.yview)
        self.income_tree.configure(yscrollcommand=scrollbar.set)
        
        self.income_tree.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')
        
        delete_button = tk.Button(
            income_frame,
            text="Удалить выбранный доход",
            command=self.delete_income,
            bg=self.colors['danger'],
            fg='white',
            font=('Arial', 10),
            padx=10,
            pady=5
        )
        delete_button.pack(pady=10)
        
    def setup_credit_tab(self):
        """Вкладка кредитов - С ВАЛИДАЦИЕЙ"""
        credit_frame = tk.Frame(self.notebook, bg=self.colors['white'])
        self.notebook.add(credit_frame, text='Кредиты')
        
        tk.Label(
            credit_frame,
            text="ВАШИ КРЕДИТЫ",
            font=('Arial', 20, 'bold'),
            fg=self.colors['primary'],
            bg=self.colors['white']
        ).pack(pady=20)
        
        form_frame = tk.Frame(credit_frame, bg='#FFEBEE', relief='ridge', borderwidth=2)
        form_frame.pack(pady=10, padx=30, fill='x')
        
        tk.Label(
            form_frame,
            text="Добавить новый кредит",
            font=('Arial', 12, 'bold'),
            fg=self.colors['danger'],
            bg='#FFEBEE'
        ).pack(pady=10)
        
        fields_frame = tk.Frame(form_frame, bg='#FFEBEE')
        fields_frame.pack(pady=10, padx=20)
        
        left_frame = tk.Frame(fields_frame, bg='#FFEBEE')
        left_frame.grid(row=0, column=0, padx=20)
        
        tk.Label(left_frame, text="Название кредита", bg='#FFEBEE').pack(anchor='w', pady=5)
        self.credit_name = tk.Entry(left_frame, width=30)
        self.credit_name.pack(pady=5)
        
        tk.Label(left_frame, text="Общая сумма (руб)", bg='#FFEBEE').pack(anchor='w', pady=5)
        self.credit_amount = tk.Entry(left_frame, width=30)
        self.credit_amount.pack(pady=5)
        
        tk.Label(left_frame, text="Годовая ставка (%)", bg='#FFEBEE').pack(anchor='w', pady=5)
        self.credit_rate = tk.Entry(left_frame, width=30)
        self.credit_rate.pack(pady=5)
        
        tk.Label(left_frame, text="Срок (месяцев)", bg='#FFEBEE').pack(anchor='w', pady=5)
        self.credit_months = tk.Entry(left_frame, width=30)
        self.credit_months.pack(pady=5)
        
        right_frame = tk.Frame(fields_frame, bg='#FFEBEE')
        right_frame.grid(row=0, column=1, padx=20)
        
        tk.Label(right_frame, text="Дата начала (ДД.ММ.ГГГГ)", bg='#FFEBEE').pack(anchor='w', pady=5)
        self.credit_start_date = tk.Entry(right_frame, width=30)
        self.credit_start_date.pack(pady=5)
        
        tk.Label(right_frame, text="Уже оплачено (руб)", bg='#FFEBEE').pack(anchor='w', pady=5)
        self.credit_paid = tk.Entry(right_frame, width=30)
        self.credit_paid.insert(0, "0")
        self.credit_paid.pack(pady=5)
        
        tk.Label(right_frame, text="Тип кредита", bg='#FFEBEE').pack(anchor='w', pady=5)
        self.credit_type = ttk.Combobox(right_frame, width=27, values=["Ипотека", "Автокредит", "Потребительский", "Кредитная карта"])
        self.credit_type.pack(pady=5)
        self.credit_type.set("Потребительский")
        
        tk.Label(right_frame, text="Банк", bg='#FFEBEE').pack(anchor='w', pady=5)
        self.credit_bank = tk.Entry(right_frame, width=30)
        self.credit_bank.pack(pady=5)
        
        add_button = tk.Button(
            form_frame,
            text="Добавить кредит",
            command=self.add_credit,
            bg=self.colors['danger'],
            fg='white',
            font=('Arial', 10, 'bold'),
            padx=20,
            pady=10
        )
        add_button.pack(pady=20)
        
        self.payment_info = tk.Label(
            form_frame,
            text="Ежемесячный платеж рассчитывается после добавления",
            font=('Arial', 10),
            fg=self.colors['primary'],
            bg='#FFEBEE'
        )
        self.payment_info.pack(pady=10)
        
        table_frame = tk.Frame(credit_frame, bg=self.colors['white'])
        table_frame.pack(pady=20, padx=30, fill='both', expand=True)
        
        tk.Label(
            table_frame,
            text="СПИСОК КРЕДИТОВ",
            font=('Arial', 14, 'bold'),
            fg=self.colors['primary'],
            bg=self.colors['white']
        ).pack(anchor='w', pady=10)
        
        columns = ('#', 'Кредит', 'Банк', 'Тип', 'Сумма', 'Ставка', 'Срок', 'Платеж')
        self.credit_tree = ttk.Treeview(table_frame, columns=columns, show='headings', height=8)
        
        col_widths = [40, 150, 100, 100, 100, 80, 80, 100]
        for idx, col in enumerate(columns):
            self.credit_tree.heading(col, text=col)
            self.credit_tree.column(col, width=col_widths[idx])
        
        scrollbar = ttk.Scrollbar(table_frame, orient='vertical', command=self.credit_tree.yview)
        self.credit_tree.configure(yscrollcommand=scrollbar.set)
        
        self.credit_tree.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')
        
        delete_button = tk.Button(
            credit_frame,
            text="Удалить выбранный кредит",
            command=self.delete_credit,
            bg=self.colors['danger'],
            fg='white',
            font=('Arial', 10),
            padx=10,
            pady=5
        )
        delete_button.pack(pady=10)
        
    def setup_expense_tab(self):
        """Вкладка расходов"""
        expense_frame = tk.Frame(self.notebook, bg=self.colors['white'])
        self.notebook.add(expense_frame, text='Расходы')
        
        tk.Label(
            expense_frame,
            text="ВАШИ РАСХОДЫ",
            font=('Arial', 20, 'bold'),
            fg=self.colors['primary'],
            bg=self.colors['white']
        ).pack(pady=20)
        
        form_frame = tk.Frame(expense_frame, bg='#FFF3E0', relief='ridge', borderwidth=2)
        form_frame.pack(pady=10, padx=30, fill='x')
        
        tk.Label(
            form_frame,
            text="Добавить новый расход",
            font=('Arial', 12, 'bold'),
            fg=self.colors['warning'],
            bg='#FFF3E0'
        ).pack(pady=10)
        
        fields_frame = tk.Frame(form_frame, bg='#FFF3E0')
        fields_frame.pack(pady=10, padx=20)
        
        tk.Label(fields_frame, text="Название расхода", bg='#FFF3E0').grid(row=0, column=0, sticky='w', pady=5)
        self.expense_name = tk.Entry(fields_frame, width=30)
        self.expense_name.grid(row=0, column=1, pady=5, padx=10)
        
        tk.Label(fields_frame, text="Сумма в месяц (руб)", bg='#FFF3E0').grid(row=1, column=0, sticky='w', pady=5)
        self.expense_amount = tk.Entry(fields_frame, width=30)
        self.expense_amount.grid(row=1, column=1, pady=5, padx=10)
        
        tk.Label(fields_frame, text="Категория", bg='#FFF3E0').grid(row=2, column=0, sticky='w', pady=5)
        self.expense_category = ttk.Combobox(fields_frame, width=28, values=["Обязательный", "Переменный", "Развлечения", "Продукты", "Транспорт"])
        self.expense_category.grid(row=2, column=1, pady=5, padx=10)
        self.expense_category.set("Обязательный")
        
        add_button = tk.Button(
            form_frame,
            text="Добавить расход",
            command=self.add_expense,
            bg=self.colors['warning'],
            fg='white',
            font=('Arial', 10, 'bold'),
            padx=20,
            pady=10
        )
        add_button.pack(pady=20)
        
        table_frame = tk.Frame(expense_frame, bg=self.colors['white'])
        table_frame.pack(pady=20, padx=30, fill='both', expand=True)
        
        tk.Label(
            table_frame,
            text="СПИСОК РАСХОДОВ",
            font=('Arial', 14, 'bold'),
            fg=self.colors['primary'],
            bg=self.colors['white']
        ).pack(anchor='w', pady=10)
        
        columns = ('#', 'Расход', 'Сумма/мес', 'Категория')
        self.expense_tree = ttk.Treeview(table_frame, columns=columns, show='headings', height=8)
        
        for col in columns:
            self.expense_tree.heading(col, text=col)
            self.expense_tree.column(col, width=150)
        
        scrollbar = ttk.Scrollbar(table_frame, orient='vertical', command=self.expense_tree.yview)
        self.expense_tree.configure(yscrollcommand=scrollbar.set)
        
        self.expense_tree.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')
        
        delete_button = tk.Button(
            expense_frame,
            text="Удалить выбранный расход",
            command=self.delete_expense,
            bg=self.colors['danger'],
            fg='white',
            font=('Arial', 10),
            padx=10,
            pady=5
        )
        delete_button.pack(pady=10)
        
    def setup_goal_tab(self):
        """Вкладка целей"""
        goal_frame = tk.Frame(self.notebook, bg=self.colors['white'])
        self.notebook.add(goal_frame, text='Цели')
        
        tk.Label(
            goal_frame,
            text="ВАШИ ФИНАНСОВЫЕ ЦЕЛИ",
            font=('Arial', 20, 'bold'),
            fg=self.colors['primary'],
            bg=self.colors['white']
        ).pack(pady=20)
        
        form_frame = tk.Frame(goal_frame, bg='#E8EAF6', relief='ridge', borderwidth=2)
        form_frame.pack(pady=10, padx=30, fill='x')
        
        tk.Label(
            form_frame,
            text="Добавить новую цель",
            font=('Arial', 12, 'bold'),
            fg=self.colors['secondary'],
            bg='#E8EAF6'
        ).pack(pady=10)
        
        fields_frame = tk.Frame(form_frame, bg='#E8EAF6')
        fields_frame.pack(pady=10, padx=20)
        
        tk.Label(fields_frame, text="Название цели", bg='#E8EAF6').grid(row=0, column=0, sticky='w', pady=5)
        self.goal_name = tk.Entry(fields_frame, width=30)
        self.goal_name.grid(row=0, column=1, pady=5, padx=10)
        
        tk.Label(fields_frame, text="Целевая сумма (руб)", bg='#E8EAF6').grid(row=1, column=0, sticky='w', pady=5)
        self.goal_target = tk.Entry(fields_frame, width=30)
        self.goal_target.grid(row=1, column=1, pady=5, padx=10)
        
        tk.Label(fields_frame, text="Уже накоплено (руб)", bg='#E8EAF6').grid(row=2, column=0, sticky='w', pady=5)
        self.goal_current = tk.Entry(fields_frame, width=30)
        self.goal_current.insert(0, "0")
        self.goal_current.grid(row=2, column=1, pady=5, padx=10)
        
        tk.Label(fields_frame, text="Срок цели (ДД.ММ.ГГГГ)", bg='#E8EAF6').grid(row=3, column=0, sticky='w', pady=5)
        self.goal_deadline = tk.Entry(fields_frame, width=30)
        self.goal_deadline.grid(row=3, column=1, pady=5, padx=10)
        
        tk.Label(fields_frame, text="Приоритет", bg='#E8EAF6').grid(row=4, column=0, sticky='w', pady=5)
        self.goal_priority = ttk.Combobox(fields_frame, width=28, values=["Высокий", "Средний", "Низкий"])
        self.goal_priority.grid(row=4, column=1, pady=5, padx=10)
        self.goal_priority.set("Высокий")
        
        add_button = tk.Button(
            form_frame,
            text="Добавить цель",
            command=self.add_goal,
            bg=self.colors['secondary'],
            fg='white',
            font=('Arial', 10, 'bold'),
            padx=20,
            pady=10
        )
        add_button.pack(pady=20)
        
        table_frame = tk.Frame(goal_frame, bg=self.colors['white'])
        table_frame.pack(pady=20, padx=30, fill='both', expand=True)
        
        tk.Label(
            table_frame,
            text="СПИСОК ЦЕЛЕЙ",
            font=('Arial', 14, 'bold'),
            fg=self.colors['primary'],
            bg=self.colors['white']
        ).pack(anchor='w', pady=10)
        
        columns = ('#', 'Цель', 'Целевая сумма', 'Накоплено', 'Осталось', 'Срок', 'Прогресс')
        self.goal_tree = ttk.Treeview(table_frame, columns=columns, show='headings', height=8)
        
        col_widths = [40, 150, 120, 120, 120, 100, 80]
        for idx, col in enumerate(columns):
            self.goal_tree.heading(col, text=col)
            self.goal_tree.column(col, width=col_widths[idx])
        
        scrollbar = ttk.Scrollbar(table_frame, orient='vertical', command=self.goal_tree.yview)
        self.goal_tree.configure(yscrollcommand=scrollbar.set)
        
        self.goal_tree.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')
        
        delete_button = tk.Button(
            goal_frame,
            text="Удалить выбранную цель",
            command=self.delete_goal,
            bg=self.colors['danger'],
            fg='white',
            font=('Arial', 10),
            padx=10,
            pady=5
        )
        delete_button.pack(pady=10)
        
    def setup_summary_tab(self):
        """Вкладка итогов"""
        summary_frame = tk.Frame(self.notebook, bg=self.colors['white'])
        self.notebook.add(summary_frame, text='Итоги')
        
        tk.Label(
            summary_frame,
            text="ФИНАНСОВАЯ СВОДКА",
            font=('Arial', 20, 'bold'),
            fg=self.colors['primary'],
            bg=self.colors['white']
        ).pack(pady=20)
        
        metrics_frame = tk.Frame(summary_frame, bg=self.colors['white'])
        metrics_frame.pack(pady=10, padx=30, fill='x')
        
        self.metric_cards = {}
        
        card1 = self.create_metric_card(metrics_frame, 0, 0, "ОБЩИЙ ДОХОД", "0 ₽/мес", "#4CAF50")
        self.metric_cards['income'] = card1
        
        card2 = self.create_metric_card(metrics_frame, 0, 1, "КРЕДИТНАЯ НАГРУЗКА", "0 ₽/мес", "#FF9800")
        self.metric_cards['credit'] = card2
        
        card3 = self.create_metric_card(metrics_frame, 1, 0, "ОБЩИЕ РАСХОДЫ", "0 ₽/мес", "#F44336")
        self.metric_cards['expense'] = card3
        
        card4 = self.create_metric_card(metrics_frame, 1, 1, "СВОБОДНЫЕ СРЕДСТВА", "0 ₽/мес", "#2196F3")
        self.metric_cards['free'] = card4
        
        update_button = tk.Button(
            summary_frame,
            text="Обновить сводку",
            command=self.update_summary,
            bg=self.colors['primary'],
            fg='white',
            font=('Arial', 12, 'bold'),
            padx=20,
            pady=10
        )
        update_button.pack(pady=20)
        
        info_text = tk.Text(
            summary_frame,
            height=8,
            width=80,
            font=('Arial', 10),
            bg='#F9F9F9',
            relief='solid',
            borderwidth=1
        )
        info_text.pack(pady=10, padx=30)
        info_text.insert('1.0', "ИНФОРМАЦИЯ\n\n")
        info_text.insert('end', "• Заполните все вкладки для получения полной финансовой картины\n")
        info_text.insert('end', "• Нажмите 'Обновить сводку' для пересчета показателей\n")
        info_text.insert('end', "• После создания файла все данные сохранятся в Excel\n")
        info_text.insert('end', "• Вы сможете редактировать данные непосредственно в Excel файле\n")
        info_text.insert('end', "• Все формулы в Excel файле динамические и обновляются автоматически")
        info_text.config(state='disabled')
        
    def create_metric_card(self, parent, row, col, title, value, color):
        """Создание карточки с метрикой"""
        card = tk.Frame(parent, bg=color, relief='raised', borderwidth=2)
        card.grid(row=row, column=col, padx=10, pady=10, sticky='nsew')
        
        title_label = tk.Label(
            card,
            text=title,
            font=('Arial', 12, 'bold'),
            fg='white',
            bg=color
        )
        title_label.pack(pady=(15, 5))
        
        value_label = tk.Label(
            card,
            text=value,
            font=('Arial', 16, 'bold'),
            fg='white',
            bg=color
        )
        value_label.pack(pady=(5, 15))
        
        card.value_label = value_label
        return card
        
    def setup_footer(self):
        """Нижняя панель с кнопками"""
        footer_frame = tk.Frame(self.root, bg=self.colors['primary'], height=80)
        footer_frame.pack(side='bottom', fill='x', padx=20, pady=10)
        
        create_button = tk.Button(
            footer_frame,
            text="СОЗДАТЬ ОТЧЕТ",
            command=self.create_report,
            bg=self.colors['success'],
            fg='white',
            font=('Arial', 14, 'bold'),
            padx=30,
            pady=15
        )
        create_button.pack(side='left', padx=20)
        
        clear_button = tk.Button(
            footer_frame,
            text="ОЧИСТИТЬ ВСЕ",
            command=self.clear_all_data,
            bg=self.colors['danger'],
            fg='white',
            font=('Arial', 12),
            padx=20,
            pady=10
        )
        clear_button.pack(side='right', padx=20)
        
        save_button = tk.Button(
            footer_frame,
            text="СОХРАНИТЬ ДАННЫЕ",
            command=self.save_data,
            bg=self.colors['secondary'],
            fg='white',
            font=('Arial', 12),
            padx=20,
            pady=10
        )
        save_button.pack(side='right', padx=10)
        
        self.status_label = tk.Label(
            footer_frame,
            text="Готов к работе",
            font=('Arial', 10),
            fg='white',
            bg=self.colors['primary']
        )
        self.status_label.pack(side='right', padx=20)
        
    def add_income(self):
        """Добавление дохода"""
        try:
            if not self.income_name.get() or not self.income_amount.get():
                messagebox.showwarning("Внимание", "Заполните название и сумму дохода!")
                return
            
            income = {
                'id': len(self.incomes) + 1,
                'name': self.income_name.get(),
                'amount': float(self.income_amount.get().replace(',', '.')),
                'category': self.income_category.get(),
                'stability': self.income_stability.get(),
                'frequency': 'monthly'
            }
            
            self.incomes.append(income)
            
            self.income_tree.insert('', 'end', values=(
                len(self.incomes),
                income['name'],
                f"{income['amount']:.0f} ₽",
                income['category'],
                income['stability']
            ))
            
            self.income_name.delete(0, 'end')
            self.income_amount.delete(0, 'end')
            
            if hasattr(self, 'status_label') and self.status_label:
                self.status_label.config(text=f"Добавлен доход: {income['name']}")
            
        except ValueError:
            messagebox.showerror("Ошибка", "Некорректная сумма дохода!")
            
    def add_credit(self):
        """Добавление кредита - С ВАЛИДАЦИЕЙ"""
        try:
            required_fields = [
                self.credit_name, self.credit_amount, self.credit_rate,
                self.credit_months, self.credit_start_date
            ]
            
            if not all(field.get() for field in required_fields):
                messagebox.showwarning("Внимание", "Заполните все обязательные поля!")
                return
            
            # ВАЛИДАЦИЯ СТАВКИ
            rate = float(self.credit_rate.get().replace(',', '.')) / 100
            if rate > 1.0:
                messagebox.showerror("Ошибка", "Ставка не может превышать 100%!")
                return
            
            # ВАЛИДАЦИЯ СРОКА
            months = int(self.credit_months.get())
            if months <= 0 or months > 600:
                messagebox.showerror("Ошибка", "Некорректный срок кредита! (1-600 месяцев)")
                return
            
            # ВАЛИДАЦИЯ ДАТЫ
            try:
                start_date = datetime.strptime(self.credit_start_date.get(), '%d.%m.%Y')
                if start_date > self.today + timedelta(days=365*5):
                    messagebox.showerror("Ошибка", "Дата начала не может быть более чем на 5 лет вперед!")
                    return
            except ValueError:
                messagebox.showerror("Ошибка", "Некорректный формат даты! Используйте ДД.ММ.ГГГГ")
                return
            
            # Рассчитываем платеж
            credit_data = {
                'amount': float(self.credit_amount.get().replace(',', '.')),
                'rate': rate,
                'months': months
            }
            
            monthly_payment = self.companion.calculate_monthly_payment(credit_data)
            
            credit = {
                'id': len(self.credits) + 1,
                'name': self.credit_name.get(),
                'amount': credit_data['amount'],
                'rate': credit_data['rate'],
                'months': credit_data['months'],
                'start_date': self.credit_start_date.get(),
                'paid': float(self.credit_paid.get().replace(',', '.')) if self.credit_paid.get() else 0,
                'type': self.credit_type.get(),
                'bank': self.credit_bank.get() if self.credit_bank.get() else "Не указан",
                'monthly_payment': monthly_payment
            }
            
            self.credits.append(credit)
            
            self.credit_tree.insert('', 'end', values=(
                len(self.credits),
                credit['name'],
                credit['bank'],
                credit['type'],
                f"{credit['amount']:.0f} ₽",
                f"{credit['rate']*100:.1f}%",
                f"{credit['months']} мес",
                f"{credit['monthly_payment']:.0f} ₽"
            ))
            
            self.payment_info.config(
                text=f"Ежемесячный платеж: {monthly_payment:.0f} ₽ | Полная стоимость: {(monthly_payment * months):.0f} ₽"
            )
            
            self.credit_name.delete(0, 'end')
            self.credit_amount.delete(0, 'end')
            self.credit_rate.delete(0, 'end')
            self.credit_months.delete(0, 'end')
            self.credit_start_date.delete(0, 'end')
            self.credit_paid.delete(0, 'end')
            self.credit_paid.insert(0, "0")
            self.credit_bank.delete(0, 'end')
            
            if hasattr(self, 'status_label') and self.status_label:
                self.status_label.config(text=f"Добавлен кредит: {credit['name']}")
            
        except ValueError as e:
            messagebox.showerror("Ошибка", f"Некорректные данные: {str(e)}")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка при добавлении кредита: {str(e)}")
            
    def add_expense(self):
        """Добавление расхода"""
        try:
            if not self.expense_name.get() or not self.expense_amount.get():
                messagebox.showwarning("Внимание", "Заполните название и сумму расхода!")
                return
            
            expense = {
                'id': len(self.expenses) + 1,
                'name': self.expense_name.get(),
                'amount': float(self.expense_amount.get().replace(',', '.')),
                'category': self.expense_category.get()
            }
            
            self.expenses.append(expense)
            
            self.expense_tree.insert('', 'end', values=(
                len(self.expenses),
                expense['name'],
                f"{expense['amount']:.0f} ₽",
                expense['category']
            ))
            
            self.expense_name.delete(0, 'end')
            self.expense_amount.delete(0, 'end')
            
            if hasattr(self, 'status_label') and self.status_label:
                self.status_label.config(text=f"Добавлен расход: {expense['name']}")
            
        except ValueError:
            messagebox.showerror("Ошибка", "Некорректная сумма расхода!")
            
    def add_goal(self):
        """Добавление цели"""
        try:
            if not self.goal_name.get() or not self.goal_target.get() or not self.goal_deadline.get():
                messagebox.showwarning("Внимание", "Заполните название, сумму и срок цели!")
                return
            
            goal = {
                'id': len(self.goals) + 1,
                'name': self.goal_name.get(),
                'target': float(self.goal_target.get().replace(',', '.')),
                'current': float(self.goal_current.get().replace(',', '.')) if self.goal_current.get() else 0,
                'deadline': self.goal_deadline.get(),
                'priority': self.goal_priority.get()
            }
            
            self.goals.append(goal)
            
            self.goal_tree.insert('', 'end', values=(
                len(self.goals),
                goal['name'],
                f"{goal['target']:.0f} ₽",
                f"{goal['current']:.0f} ₽",
                f"{goal['target'] - goal['current']:.0f} ₽",
                goal['deadline'],
                f"{(goal['current']/goal['target']*100):.1f}%" if goal['target'] > 0 else "0%"
            ))
            
            self.goal_name.delete(0, 'end')
            self.goal_target.delete(0, 'end')
            self.goal_current.delete(0, 'end')
            self.goal_current.insert(0, "0")
            self.goal_deadline.delete(0, 'end')
            
            if hasattr(self, 'status_label') and self.status_label:
                self.status_label.config(text=f"Добавлена цель: {goal['name']}")
            
        except ValueError:
            messagebox.showerror("Ошибка", "Некорректные данные цели!")
            
    def delete_income(self):
        """Удаление дохода"""
        selected = self.income_tree.selection()
        if not selected:
            messagebox.showwarning("Внимание", "Выберите доход для удаления!")
            return
        
        if not messagebox.askyesno("Подтверждение", "Удалить выбранный доход?"):
            return
        
        for item in selected:
            self.income_tree.delete(item)
        
        self.incomes = []
        for item in self.income_tree.get_children():
            values = self.income_tree.item(item)['values']
            if values:
                self.incomes.append({
                    'id': int(values[0]),
                    'name': values[1],
                    'amount': float(values[2].replace(' ₽', '').replace(',', '')),
                    'category': values[3],
                    'stability': values[4]
                })
        
        if hasattr(self, 'status_label') and self.status_label:
            self.status_label.config(text="Доход удален")
        self.update_summary()
        
    def delete_credit(self):
        """Удаление кредита"""
        selected = self.credit_tree.selection()
        if not selected:
            messagebox.showwarning("Внимание", "Выберите кредит для удаления!")
            return
        
        if not messagebox.askyesno("Подтверждение", "Удалить выбранный кредит?"):
            return
        
        for item in selected:
            self.credit_tree.delete(item)
        
        self.credits = []
        for item in self.credit_tree.get_children():
            values = self.credit_tree.item(item)['values']
            if values:
                self.credits.append({
                    'id': int(values[0]),
                    'name': values[1],
                    'bank': values[2],
                    'type': values[3],
                    'amount': float(values[4].replace(' ₽', '').replace(',', '')),
                    'rate': float(values[5].replace('%', '')) / 100,
                    'months': int(values[6].replace(' мес', '')),
                    'monthly_payment': float(values[7].replace(' ₽', '').replace(',', ''))
                })
        
        if hasattr(self, 'status_label') and self.status_label:
            self.status_label.config(text="Кредит удален")
        self.update_summary()
        
    def delete_expense(self):
        """Удаление расхода"""
        selected = self.expense_tree.selection()
        if not selected:
            messagebox.showwarning("Внимание", "Выберите расход для удаления!")
            return
        
        if not messagebox.askyesno("Подтверждение", "Удалить выбранный расход?"):
            return
        
        for item in selected:
            self.expense_tree.delete(item)
        
        self.expenses = []
        for item in self.expense_tree.get_children():
            values = self.expense_tree.item(item)['values']
            if values:
                self.expenses.append({
                    'id': int(values[0]),
                    'name': values[1],
                    'amount': float(values[2].replace(' ₽', '').replace(',', '')),
                    'category': values[3]
                })
        
        if hasattr(self, 'status_label') and self.status_label:
            self.status_label.config(text="Расход удален")
        self.update_summary()
        
    def delete_goal(self):
        """Удаление цели"""
        selected = self.goal_tree.selection()
        if not selected:
            messagebox.showwarning("Внимание", "Выберите цель для удаления!")
            return
        
        if not messagebox.askyesno("Подтверждение", "Удалить выбранную цель?"):
            return
        
        for item in selected:
            self.goal_tree.delete(item)
        
        self.goals = []
        for item in self.goal_tree.get_children():
            values = self.goal_tree.item(item)['values']
            if values:
                self.goals.append({
                    'id': int(values[0]),
                    'name': values[1],
                    'target': float(values[2].replace(' ₽', '').replace(',', '')),
                    'current': float(values[3].replace(' ₽', '').replace(',', '')),
                    'priority': "Высокий"
                })
        
        if hasattr(self, 'status_label') and self.status_label:
            self.status_label.config(text="Цель удалена")
        self.update_summary()
        
    def update_summary(self):
        """Обновление сводки"""
        try:
            total_income = sum(income['amount'] for income in self.incomes)
            total_credit_payments = sum(credit.get('monthly_payment', 0) for credit in self.credits)
            total_expenses = sum(expense['amount'] for expense in self.expenses)
            free_money = total_income - total_credit_payments - total_expenses
            
            self.metric_cards['income'].value_label.config(text=f"{total_income:.0f} ₽/мес")
            self.metric_cards['credit'].value_label.config(text=f"{total_credit_payments:.0f} ₽/мес")
            self.metric_cards['expense'].value_label.config(text=f"{total_expenses:.0f} ₽/мес")
            self.metric_cards['free'].value_label.config(text=f"{free_money:.0f} ₽/мес")
            
            if hasattr(self, 'status_label') and self.status_label:
                self.status_label.config(text="Сводка обновлена")
            
        except Exception as e:
            if hasattr(self, 'status_label') and self.status_label:
                self.status_label.config(text="Ошибка обновления сводки")
            
    def create_report(self):
        """Создание отчета"""
        try:
            if not self.incomes and not self.credits:
                messagebox.showwarning("Внимание", "Добавьте хотя бы один доход или кредит!")
                return
            
            self.companion.incomes = self.incomes
            self.companion.credits = self.credits
            self.companion.expenses = self.expenses
            self.companion.goals = self.goals
            
            filename = self.companion.create_excel_file()
            
            if filename:
                os.startfile(os.path.dirname(os.path.abspath(filename)))
                
                messagebox.showinfo(
                    "Успех!",
                    f"Отчет успешно создан!\n\n"
                    f"Файл: {filename}\n\n"
                    f"В файле созданы:\n"
                    f"• Динамический дашборд с метриками\n"
                    f"• Профессиональные диаграммы\n"
                    f"• Интерактивные спидометры\n"
                    f"• Умные уведомления\n"
                    f"• Полная аналитика\n"
                    f"• Анализ рисков\n\n"
                    f"Открыта папка с файлом!"
                )
                
                if hasattr(self, 'status_label') and self.status_label:
                    self.status_label.config(text=f"Отчет создан: {os.path.basename(filename)}")
            else:
                messagebox.showerror("Ошибка", "Не удалось создать отчет!")
                if hasattr(self, 'status_label') and self.status_label:
                    self.status_label.config(text="Ошибка создания отчета")
                
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка при создании отчета\n{str(e)}")
            if hasattr(self, 'status_label') and self.status_label:
                self.status_label.config(text="Ошибка создания отчета")
                
    def clear_all_data(self):
        """Очистка всех данных"""
        if not messagebox.askyesno("Подтверждение", "Очистить ВСЕ данные? Это действие нельзя отменить."):
            return
        
        self.incomes = []
        self.credits = []
        self.expenses = []
        self.goals = []
        
        for tree in [self.income_tree, self.credit_tree, self.expense_tree, self.goal_tree]:
            for item in tree.get_children():
                tree.delete(item)
        
        self.update_summary()
        
        if hasattr(self, 'status_label') and self.status_label:
            self.status_label.config(text="Все данные очищены")
        
    def save_data(self):
        """Сохранение данных в файл"""
        try:
            data = {
                'incomes': self.incomes,
                'credits': self.credits,
                'expenses': self.expenses,
                'goals': self.goals
            }
            
            filename = f"финансовые_данные_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
            
            with open(filename, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2, default=str)
            
            if hasattr(self, 'status_label') and self.status_label:
                self.status_label.config(text=f"Данные сохранены в {filename}")
            
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка при сохранении данных\n{str(e)}")
            
    def run(self):
        """Запуск приложения"""
        self.root.mainloop()


def main():
    """Главная функция с GUI"""
    print("\n" + "*"*70)
    print("        ДОБРО ПОЖАЛОВАТЬ В ПРОФЕССИОНАЛЬНЫЙ ФИНАНСОВЫЙ КОМПАНЬОН 2026        ")
    print("*"*70)
    
    app = FinancialCompanionGUI()
    app.run()

if __name__ == "__main__":
    main()